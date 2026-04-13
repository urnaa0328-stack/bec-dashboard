import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook


# =========================================================
# HELPERS
# =========================================================
def _normalize_sheet_name(name: str) -> str:
    return (
        str(name)
        .strip()
        .lower()
        .replace("_", "")
        .replace("-", "")
        .replace(" ", "")
    )


def _find_sheet_name(excel_path: str, candidates: list[str]) -> tuple[str | None, list[str]]:
    try:
        xls = pd.ExcelFile(excel_path)
        actual_sheets = xls.sheet_names
    except Exception as e:
        st.error(f"Excel файл нээхэд алдаа гарлаа: {e}")
        return None, []

    normalized_map = {_normalize_sheet_name(s): s for s in actual_sheets}

    for candidate in candidates:
        key = _normalize_sheet_name(candidate)
        if key in normalized_map:
            return normalized_map[key], actual_sheets

    return None, actual_sheets


def _read_sheet_auto(excel_path: str, candidates: list[str], label: str) -> tuple[pd.DataFrame | None, str | None]:
    matched_sheet, actual_sheets = _find_sheet_name(excel_path, candidates)

    if matched_sheet is None:
        return None, None

    try:
        df = pd.read_excel(excel_path, sheet_name=matched_sheet)
        df.columns = [str(c).strip() for c in df.columns]
        return df, matched_sheet
    except Exception as e:
        st.error(f"{label} sheet уншихад алдаа гарлаа: {e}")
        return None, None


def _ensure_text(df: pd.DataFrame, cols: list[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c not in df.columns:
            df[c] = ""
        df[c] = df[c].fillna("").astype(str).str.strip()
    return df


def _name_match(task_owner: str, employee_name: str) -> bool:
    task_owner = str(task_owner).strip().lower()
    employee_name = str(employee_name).strip().lower()

    if not task_owner or not employee_name:
        return False

    if task_owner == employee_name:
        return True

    normalized = (
        task_owner.replace(";", ",")
        .replace("/", ",")
        .replace("\n", ",")
    )
    parts = [x.strip() for x in normalized.split(",") if x.strip()]

    if employee_name in parts:
        return True

    return employee_name in task_owner


def _is_done_status(text: str) -> bool:
    txt = str(text).strip().lower()
    done_words = ["дууссан", "хийгдсэн", "closed", "completed", "done", "resolved"]
    return any(word in txt for word in done_words)


# =========================================================
# TASK LOADERS
# =========================================================
def _load_operation_sheet(excel_path: str, system_name: str, candidates: list[str]) -> pd.DataFrame:
    df, matched_sheet = _read_sheet_auto(excel_path, candidates, system_name)
    if df is None or matched_sheet is None:
        return pd.DataFrame()

    df = _ensure_text(df, ["Ажлын төрөл", "Төслийн нэр", "Явц", "Хариуцагч", "Дэмжлэг"])

    df = df[
        (df["Төслийн нэр"] != "") |
        (df["Ажлын төрөл"] != "") |
        (df["Хариуцагч"] != "")
    ].copy()

    if df.empty:
        return pd.DataFrame()

    df = df[~df["Явц"].apply(_is_done_status)].copy()

    df["_task_source"] = "operation"
    df["_sheet_name"] = matched_sheet
    df["_system"] = system_name
    df["_owner"] = df["Хариуцагч"]
    df["_task_name"] = df["Төслийн нэр"]
    df["_task_type"] = df["Ажлын төрөл"]
    df["_status"] = df["Явц"]
    df["_match_key1"] = df["Төслийн нэр"]
    df["_match_key2"] = df["Ажлын төрөл"]

    df["task_label"] = (
        df["_system"].astype(str)
        + " | "
        + df["_task_name"].astype(str)
        + " | "
        + df["_task_type"].astype(str)
        + " | "
        + df["_status"].astype(str)
    )

    return df


def _load_allmed_ticket_sheet(excel_path: str) -> pd.DataFrame:
    df, matched_sheet = _read_sheet_auto(
        excel_path,
        candidates=[
            "AllMed Ticket",
            "Ticket AllMed",
            "AllMedTicket",
            "AllMed",
            "ticket allmed",
        ],
        label="AllMed Ticket",
    )

    if df is None or matched_sheet is None:
        return pd.DataFrame()

    df = _ensure_text(df, ["Явц", "Ажлын нэр", "Хариуцагчийн тайлбар", "Хариуцагч"])

    df = df[
        (df["Ажлын нэр"] != "") |
        (df["Хариуцагч"] != "")
    ].copy()

    if df.empty:
        return pd.DataFrame()

    df = df[~df["Явц"].apply(_is_done_status)].copy()

    df["_task_source"] = "allmed_ticket"
    df["_sheet_name"] = matched_sheet
    df["_system"] = "AllMed Ticket"
    df["_owner"] = df["Хариуцагч"]
    df["_task_name"] = df["Ажлын нэр"]
    df["_task_type"] = "Ticket"
    df["_status"] = df["Явц"]
    df["_match_key1"] = df["Ажлын нэр"]
    df["_match_key2"] = df["Хариуцагч"]

    df["task_label"] = (
        df["_system"].astype(str)
        + " | "
        + df["_task_name"].astype(str)
        + " | "
        + df["_status"].astype(str)
        + " | "
        + df["_owner"].astype(str)
    )

    return df


def _load_all_tasks(excel_path: str) -> pd.DataFrame:
    frames = []

    allcall_op = _load_operation_sheet(
        excel_path,
        system_name="AllCall Operation",
        candidates=[
            "AllCall operation",
            "AllCall Operation",
            "Operation AllCall",
            "operation allcall",
        ],
    )
    if not allcall_op.empty:
        frames.append(allcall_op)

    allmed_op = _load_operation_sheet(
        excel_path,
        system_name="AllMed Operation",
        candidates=[
            "AllMed operation",
            "AllMed Operation",
            "Operation AllMed",
            "operation allmed",
        ],
    )
    if not allmed_op.empty:
        frames.append(allmed_op)

    allmed_ticket = _load_allmed_ticket_sheet(excel_path)
    if not allmed_ticket.empty:
        frames.append(allmed_ticket)

    if not frames:
        return pd.DataFrame()

    return pd.concat(frames, ignore_index=True)


def _get_employee_list(tasks_df: pd.DataFrame) -> list[str]:
    if tasks_df.empty or "_owner" not in tasks_df.columns:
        return []

    names = set()

    for raw in tasks_df["_owner"].fillna("").astype(str):
        txt = raw.strip()
        if not txt:
            continue

        normalized = (
            txt.replace(";", ",")
            .replace("/", ",")
            .replace("\n", ",")
        )
        parts = [x.strip() for x in normalized.split(",") if x.strip()]
        if parts:
            for p in parts:
                names.add(p)
        else:
            names.add(txt)

    return sorted(names)


def _filter_tasks_by_employee(tasks_df: pd.DataFrame, employee_name: str) -> pd.DataFrame:
    if tasks_df.empty or not employee_name:
        return pd.DataFrame()

    fdf = tasks_df.copy()
    fdf = fdf[fdf["_owner"].apply(lambda x: _name_match(x, employee_name))].copy()
    return fdf


# =========================================================
# ATTENDANCE LOG
# =========================================================
def _ensure_attendance_log_sheet(excel_path: str, sheet_name: str = "Attendance Log"):
    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        ws.append([
            "ID",
            "Ажилтан",
            "Task Source",
            "Систем",
            "Sheet",
            "Task Name",
            "Task Type",
            "Match Key 1",
            "Match Key 2",
            "Эхлэх огноо",
            "Эхлэх цаг",
            "Дуусах огноо",
            "Дуусах цаг",
            "Ажилласан минут",
            "Төлөв",
            "Тайлбар",
        ])
        wb.save(excel_path)

    wb.close()


def _read_attendance_log(excel_path: str, sheet_name: str = "Attendance Log") -> pd.DataFrame:
    _ensure_attendance_log_sheet(excel_path, sheet_name)

    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
        df.columns = [str(c).strip() for c in df.columns]
        return df
    except Exception:
        return pd.DataFrame(columns=[
            "ID",
            "Ажилтан",
            "Task Source",
            "Систем",
            "Sheet",
            "Task Name",
            "Task Type",
            "Match Key 1",
            "Match Key 2",
            "Эхлэх огноо",
            "Эхлэх цаг",
            "Дуусах огноо",
            "Дуусах цаг",
            "Ажилласан минут",
            "Төлөв",
            "Тайлбар",
        ])


def _append_attendance_log(
    excel_path: str,
    employee: str,
    task_source: str,
    system_name: str,
    sheet_name: str,
    task_name: str,
    task_type: str,
    match_key1: str,
    match_key2: str,
    note: str = "",
):
    _ensure_attendance_log_sheet(excel_path, "Attendance Log")

    wb = load_workbook(excel_path)
    ws = wb["Attendance Log"]

    now = datetime.now()
    next_id = ws.max_row

    ws.append([
        next_id,
        employee,
        task_source,
        system_name,
        sheet_name,
        task_name,
        task_type,
        match_key1,
        match_key2,
        now.strftime("%Y-%m-%d"),
        now.strftime("%H:%M:%S"),
        "",
        "",
        "",
        "Идэвхтэй",
        note,
    ])

    wb.save(excel_path)
    wb.close()


def _finish_attendance_log(excel_path: str, employee: str):
    _ensure_attendance_log_sheet(excel_path, "Attendance Log")

    wb = load_workbook(excel_path)
    ws = wb["Attendance Log"]

    now = datetime.now()
    found_row = None

    for row in range(ws.max_row, 1, -1):
        emp = ws.cell(row=row, column=2).value
        status = ws.cell(row=row, column=15).value

        if str(emp).strip() == employee and str(status).strip() == "Идэвхтэй":
            found_row = row
            break

    if found_row is None:
        wb.close()
        return False, "Идэвхтэй task олдсонгүй."

    start_date = ws.cell(found_row, 10).value
    start_time = ws.cell(found_row, 11).value

    try:
        start_dt = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M:%S")
        worked_minutes = int((now - start_dt).total_seconds() / 60)
    except Exception:
        worked_minutes = ""

    ws.cell(found_row, 12).value = now.strftime("%Y-%m-%d")
    ws.cell(found_row, 13).value = now.strftime("%H:%M:%S")
    ws.cell(found_row, 14).value = worked_minutes
    ws.cell(found_row, 15).value = "Дууссан"

    result = {
        "task_source": ws.cell(found_row, 3).value,
        "system": ws.cell(found_row, 4).value,
        "sheet": ws.cell(found_row, 5).value,
        "task_name": ws.cell(found_row, 6).value,
        "task_type": ws.cell(found_row, 7).value,
        "match_key1": ws.cell(found_row, 8).value,
        "match_key2": ws.cell(found_row, 9).value,
        "worked_minutes": worked_minutes,
    }

    wb.save(excel_path)
    wb.close()

    return True, result


def _get_active_task_for_employee(log_df: pd.DataFrame, employee: str) -> pd.DataFrame:
    if log_df.empty or not employee:
        return pd.DataFrame()

    fdf = log_df.copy()

    if "Ажилтан" not in fdf.columns or "Төлөв" not in fdf.columns:
        return pd.DataFrame()

    fdf["Ажилтан"] = fdf["Ажилтан"].fillna("").astype(str).str.strip()
    fdf["Төлөв"] = fdf["Төлөв"].fillna("").astype(str).str.strip()

    fdf = fdf[(fdf["Ажилтан"] == employee) & (fdf["Төлөв"] == "Идэвхтэй")]
    return fdf.tail(1)


# =========================================================
# UPDATE SOURCE SHEET
# =========================================================
def _update_operation_status(
    excel_path: str,
    sheet_name: str,
    project_name: str,
    task_type: str,
    new_status: str,
):
    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return False, f"{sheet_name} sheet олдсонгүй."

    ws = wb[sheet_name]

    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[str(ws.cell(1, col).value).strip()] = col

    project_col = headers.get("Төслийн нэр")
    type_col = headers.get("Ажлын төрөл")
    progress_col = headers.get("Явц")

    if not project_col or not type_col or not progress_col:
        wb.close()
        return False, "Operation sheet-д шаардлагатай багана олдсонгүй."

    target_row = None
    for row in range(2, ws.max_row + 1):
        p = str(ws.cell(row, project_col).value).strip()
        t = str(ws.cell(row, type_col).value).strip()
        if p == str(project_name).strip() and t == str(task_type).strip():
            target_row = row
            break

    if target_row is None:
        wb.close()
        return False, "Operation task мөр олдсонгүй."

    ws.cell(target_row, progress_col).value = new_status
    wb.save(excel_path)
    wb.close()

    return True, "Амжилттай шинэчлэгдлээ."


def _update_allmed_ticket_status(
    excel_path: str,
    sheet_name: str,
    task_name: str,
    owner_name: str,
    new_status: str,
):
    wb = load_workbook(excel_path)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return False, f"{sheet_name} sheet олдсонгүй."

    ws = wb[sheet_name]

    headers = {}
    for col in range(1, ws.max_column + 1):
        headers[str(ws.cell(1, col).value).strip()] = col

    task_col = headers.get("Ажлын нэр")
    owner_col = headers.get("Хариуцагч")
    progress_col = headers.get("Явц")
    done_date_col = headers.get("Дууссан огноо")

    if not task_col or not owner_col or not progress_col:
        wb.close()
        return False, "AllMed Ticket sheet-д шаардлагатай багана олдсонгүй."

    target_row = None
    for row in range(2, ws.max_row + 1):
        t = str(ws.cell(row, task_col).value).strip()
        o = str(ws.cell(row, owner_col).value).strip()

        if t == str(task_name).strip() and _name_match(o, str(owner_name).strip()):
            target_row = row
            break

    if target_row is None:
        wb.close()
        return False, "AllMed Ticket мөр олдсонгүй."

    ws.cell(target_row, progress_col).value = new_status

    if new_status == "Дууссан" and done_date_col:
        ws.cell(target_row, done_date_col).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    wb.save(excel_path)
    wb.close()

    return True, "Амжилттай шинэчлэгдлээ."


def _update_source_status(
    excel_path: str,
    task_source: str,
    sheet_name: str,
    match_key1: str,
    match_key2: str,
    new_status: str,
):
    if task_source == "operation":
        return _update_operation_status(
            excel_path=excel_path,
            sheet_name=sheet_name,
            project_name=match_key1,
            task_type=match_key2,
            new_status=new_status,
        )

    if task_source == "allmed_ticket":
        return _update_allmed_ticket_status(
            excel_path=excel_path,
            sheet_name=sheet_name,
            task_name=match_key1,
            owner_name=match_key2,
            new_status=new_status,
        )

    return False, f"Танигдаагүй task source: {task_source}"


# =========================================================
# MAIN
# =========================================================
def render_attendance_dashboard(excel_path: str):
    tasks_df = _load_all_tasks(excel_path)
    log_df = _read_attendance_log(excel_path)

    if tasks_df.empty:
        st.warning("Task олдсонгүй. AllCall operation, AllMed operation, AllMed Ticket sheet-үүдээ шалгана уу.")
        return

    st.markdown("### Clock In / Clock Out")

    c1, c2 = st.columns(2)

    employee_options = _get_employee_list(tasks_df)

    with c1:
        selected_employee = st.selectbox(
            "Ажилтан",
            [""] + employee_options,
            key="att_employee_select"
        )

    with c2:
        note = st.text_input("Тайлбар", key="att_note").strip()

    employee_tasks = _filter_tasks_by_employee(tasks_df, selected_employee)

    st.markdown("### Task")

    if selected_employee:
        if employee_tasks.empty:
            st.info("Энэ ажилтанд оноогдсон task алга.")
        else:
            selected_task_label = st.selectbox(
                "Оноогдсон task",
                employee_tasks["task_label"].tolist(),
                key="att_task_select"
            )

            selected_row = employee_tasks[employee_tasks["task_label"] == selected_task_label].iloc[0]

            st.write(f"**Эх сурвалж:** {selected_row['_system']}")
            st.write(f"**Sheet:** {selected_row['_sheet_name']}")
            st.write(f"**Task:** {selected_row['_task_name']}")
            st.write(f"**Төрөл:** {selected_row['_task_type']}")
            st.write(f"**Хариуцагч:** {selected_row['_owner']}")
            st.write(f"**Одоогийн явц:** {selected_row['_status']}")

            active_df = _get_active_task_for_employee(log_df, selected_employee)

            b1, b2 = st.columns(2)

            with b1:
                if st.button("⏱️ Clock In", use_container_width=True):
                    if active_df.empty:
                        _append_attendance_log(
                            excel_path=excel_path,
                            employee=selected_employee,
                            task_source=selected_row["_task_source"],
                            system_name=selected_row["_system"],
                            sheet_name=selected_row["_sheet_name"],
                            task_name=selected_row["_task_name"],
                            task_type=selected_row["_task_type"],
                            match_key1=selected_row["_match_key1"],
                            match_key2=selected_row["_match_key2"],
                            note=note,
                        )

                        ok, msg = _update_source_status(
                            excel_path=excel_path,
                            task_source=selected_row["_task_source"],
                            sheet_name=selected_row["_sheet_name"],
                            match_key1=selected_row["_match_key1"],
                            match_key2=selected_row["_match_key2"],
                            new_status="Хийгдэж байна",
                        )

                        if ok:
                            st.success("Clock In амжилттай. Task эхэллээ.")
                        else:
                            st.warning(f"Clock In хийгдсэн ч source sheet update дутуу: {msg}")

                        st.rerun()
                    else:
                        st.error("Энэ ажилтан аль хэдийн идэвхтэй task дээр ажиллаж байна.")

            with b2:
                if st.button("✅ Clock Out", use_container_width=True):
                    success, result = _finish_attendance_log(excel_path, selected_employee)

                    if not success:
                        st.error(result)
                    else:
                        ok, msg = _update_source_status(
                            excel_path=excel_path,
                            task_source=result["task_source"],
                            sheet_name=result["sheet"],
                            match_key1=result["match_key1"],
                            match_key2=result["match_key2"],
                            new_status="Дууссан",
                        )

                        if ok:
                            st.success(
                                f"Clock Out амжилттай. Ажилласан хугацаа: {result['worked_minutes']} минут."
                            )
                        else:
                            st.warning(
                                f"Clock Out хийгдсэн ч source sheet update дутуу: {msg}"
                            )

                        st.rerun()
    else:
        st.info("Эхлээд ажилтан сонгоно уу.")

    st.markdown("---")
    st.markdown("### Идэвхтэй task")

    if selected_employee:
        active_df = _get_active_task_for_employee(log_df, selected_employee)
        if not active_df.empty:
            st.dataframe(active_df, use_container_width=True, hide_index=True)
        else:
            st.info("Идэвхтэй task алга.")

    st.markdown("### Attendance Log")
    if not log_df.empty:
        st.dataframe(log_df.sort_index(ascending=False), use_container_width=True, hide_index=True)
    else:
        st.info("Attendance log хоосон байна.")
